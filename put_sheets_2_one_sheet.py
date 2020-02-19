import openpyxl

new_wb = openpyxl.Workbook()
dest_filename = 'output.xlsx'
ws1 = new_wb.active
ws1.title = "Data"
list = ["Дата", "Приход", "Расход", "Остаток", "Номенклатура", "Склад", "Период пополнения", "МинПартия", "МинТранспПартия", "ЦенаЗакупки", "ЦенаПродажи"]
for ind in range(1,12):
    ws1.cell(column=ind, row=1).value = list[ind-1]

wb = openpyxl.load_workbook(filename = 'input.xlsx', data_only=True)
sheet_names = wb.sheetnames
last_row = 2
for i in sheet_names:
    sheet = wb[i]  
    
    max_col = sheet.max_column+1
    max_row = sheet.max_row+1
    
    for row in range(2, max_row):
        for col in range(1,max_col):
            ws1.cell(column=col, row=last_row).value = sheet.cell(column=col, row=row).value
        last_row = last_row+1

new_wb.save(filename=dest_filename)

