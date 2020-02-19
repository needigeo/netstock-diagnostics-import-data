import openpyxl, os

new_wb = openpyxl.Workbook()
dest_filename = 'output.xlsx'
ws1 = new_wb.active
ws1.title = "Data"
list = ["Дата", "Приход", "Расход", "Остаток", "Номенклатура", "Склад", "Период пополнения", "МинПартия", "МинТранспПартия", "ЦенаЗакупки", "ЦенаПродажи"]
for ind in range(1,11):
    ws1.cell(column=ind, row=1).value = list[ind-1]

last_row = 2
for file in os.listdir():
    if not file.endswith("xlsx"):
        continue
    wb = openpyxl.load_workbook(filename = file)
    sheet_names = wb.sheetnames
    warehouse = ""
    nomenclature = ""

    for i in sheet_names:
        sheet = wb[i]  
        
        max_col = sheet.max_column+1
        max_row = sheet.max_row+1
        for row in range(2, max_row):
            for col in range(1,max_col):
                if sheet.cell(column=2, row=row).alignment.indent==0:
                    warehouse = sheet.cell(column=2, row=row).value
                    continue 
                if sheet.cell(column=2, row=row).alignment.indent==1:
                    nomenclature = sheet.cell(column=2, row=row).value
                    continue
                ws1.cell(column=col+3, row=last_row).value = sheet.cell(column=col, row=row).value
 
              
                ws1.cell(column=1, row=last_row).value  = nomenclature
                ws1.cell(column=2, row=last_row).value = warehouse
            last_row = last_row+1

new_wb.save(filename=dest_filename)

